import os

from openpyxl import Workbook
import openpyxl
from sqlalchemy.orm import sessionmaker

from schema import *
from cfg import *
from eveimport import *


def zero2none(val):
    if (val == 0):
        return None
    return val


def get_new_session():
    smaker = sessionmaker(bind=engine)
    return smaker()


def write_excel(workbook, worksheet, arr):
    xaxis = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    wb = None
    try:
        wb = openpyxl.load_workbook(workbook)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove_sheet(wb.active)
    try:
        ws = wb.get_sheet_by_name(worksheet)
        wb.remove_sheet(ws)
    except KeyError:
        pass

    ws = wb.create_sheet()
    ws.title = worksheet

    y = 1
    for row in arr:
        x = 0
        for col in row:
            ws["%s%s" % (xaxis[x], y)] = col
            x += 1
        y += 1

    wb.save(workbook)


def create_tables(engine):
    Base.metadata.create_all(engine)


def init_db(engine, session):
    return

    drop_tables = False
    if (drop_tables):
        Drop.__table__.drop(engine)
        Attacker.__table__.drop(engine)
        PriceHistory.__table__.drop(engine)
        PriceCurrent.__table__.drop(engine)
        KillMail.__table__.drop(engine)
        EveType.__table__.drop(engine)
        EveGroup.__table__.drop(engine)
        EveCategory.__table__.drop(engine)

        Alliance.__table__.drop(engine)
        Corporation.__table__.drop(engine)

        SolarSystem.__table__.drop(engine)
        Constellation.__table__.drop(engine)
        Region.__table__.drop(engine)

    create_tables(engine)

    import_regions(session, 'import/regions.csv')
    import_constellations(session, 'import/constellations.csv')
    import_systems(session, 'import/solarsystems.csv')
    import_categoryids(session, 'import/categoryIDs.yaml')
    import_groupids(session, 'import/groupIDs.yaml')
    import_typeids(session, 'import/typeIDs.yaml')
